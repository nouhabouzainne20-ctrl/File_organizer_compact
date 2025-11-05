#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Script compact de gestion de fichiers avec Excel et IFC (<200 lignes)"""
import os
import shutil
import re
import configparser
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

try:
    import ifcopenshell
    import ifcopenshell.util.element
    IFC_AVAILABLE = True
except ImportError:
    IFC_AVAILABLE = False

class FileOrganizer:
    def __init__(self, config_file='config.ini'):
        self.config = self._load_config(config_file)
        self.source = self.config.get('Paths', 'source_folder')
        self.dest = self.config.get('Paths', 'destination_base')
        self.excel_path = self.config.get('Paths', 'excel_config_file')
        self.ifc_folder = self.config.get('Paths', 'ifc_analysis_folder')
        self.analyze_ifc_enabled = self.config.get('Settings', 'analyze_ifc', fallback='yes').lower() == 'yes'
        self.mapping = {}
        self.processed = set()

    def _load_config(self, config_file):
        config = configparser.ConfigParser()
        if os.path.exists(config_file):
            config.read(config_file, encoding='utf-8')
        else:
            config['Paths'] = {
                'source_folder': str(Path.home() / 'Downloads'),
                'destination_base': str(Path.home() / 'Documents' / 'Organised_Files'),
                'excel_config_file': 'file_mapping.xlsx',
                'ifc_analysis_folder': str(Path.home() / 'Documents' / 'IFC_Analysis')}
            config['Settings'] = {'analyze_ifc': 'yes'}
            with open(config_file, 'w', encoding='utf-8') as f:
                config.write(f)
            print(f"✓ Fichier de configuration créé: {config_file}")
        return config

    def create_template(self):
        wb = Workbook()
        wb.remove(wb['Sheet'])
        ws = wb.create_sheet("google")
        ws.append(["Nom du fichier", "Sous-répertoire destination"])
        ws.append(["google.design.*.aps", "Design/Plans"])
        ws.append(["google.*.pdf", "Documents"])
        ws = wb.create_sheet("ifc")
        ws.append(["Nom du fichier", "Sous-répertoire destination"])
        ws.append(["*.ifc", "BIM/Models"])
        ws.append(["building.*.ifc", "BIM/Buildings"])
        wb.save(self.excel_path)
        print(f"✓ Template créé: {self.excel_path}")
    def load_config(self):
        if not os.path.exists(self.excel_path):
            self.create_template()
            return
        wb = load_workbook(self.excel_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rules = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    pattern = str(row[0]).strip()
                    regex = re.compile('^' + re.escape(pattern).replace(r'\*', '.*') + '$', re.IGNORECASE)
                    rules.append({'pattern': pattern, 'dest': str(row[1]).strip(), 'regex': regex})
            if rules:
                self.mapping[sheet.lower()] = rules
        print(f"✓ Config chargée: {len(self.mapping)} catégorie(s)")

    def find_destination(self, filename):
        for category, rules in self.mapping.items():
            for rule in rules:
                if rule['regex'].match(filename):
                    return category, rule['dest']
        return None, None

    def analyze_ifc(self, ifc_path):
        if not IFC_AVAILABLE:
            return None
        try:
            print(f"  📊 Analyse IFC...")
            ifc = ifcopenshell.open(ifc_path)
            data = []
            types = ['IfcWall', 'IfcSlab', 'IfcBeam', 'IfcColumn', 'IfcWindow', 'IfcDoor']
            for ifc_type in types:
                for elem in ifc.by_type(ifc_type):
                    info = {'Type': elem.is_a(), 'GlobalId': getattr(elem, 'GlobalId', 'N/A'),
                           'Name': getattr(elem, 'Name', 'Sans nom') or 'Sans nom', 'Properties': {}}
                    try:
                        psets = ifcopenshell.util.element.get_psets(elem)
                        for pset_name, pset_data in psets.items():
                            for key in ['Width', 'Height', 'Length']:
                                if key in pset_data:
                                    info['Properties'][key] = round(float(pset_data[key]), 3)
                    except:
                        pass
                    data.append(info)
            if not data:
                return None
            os.makedirs(self.ifc_folder, exist_ok=True)
            base = os.path.splitext(os.path.basename(ifc_path))[0]
            excel_file = os.path.join(self.ifc_folder, f"{base}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "IFC Analysis"
            headers = ['Type', 'GlobalId', 'Nom', 'Largeur', 'Hauteur', 'Longueur']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            for row, elem in enumerate(data, 2):
                ws.cell(row, 1, elem['Type'])
                ws.cell(row, 2, elem['GlobalId'])
                ws.cell(row, 3, elem['Name'])
                ws.cell(row, 4, elem['Properties'].get('Width'))
                ws.cell(row, 5, elem['Properties'].get('Height'))
                ws.cell(row, 6, elem['Properties'].get('Length'))
            ws_sum = wb.create_sheet("Résumé")
            ws_sum['A1'] = "Fichier analysé:"
            ws_sum['B1'] = os.path.basename(ifc_path)
            ws_sum['A2'] = "Date:"
            ws_sum['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws_sum['A3'] = "Éléments:"
            ws_sum['B3'] = len(data)
            wb.save(excel_file)
            print(f"  ✓ Analyse exportée: {len(data)} éléments")
            return excel_file
        except Exception as e:
            print(f"  ✗ Erreur IFC: {e}")
            return None

    def get_versioned_path(self, folder, name, ext):
        path = os.path.join(folder, f"{name}{ext}")
        if not os.path.exists(path):
            return path
        version = 2
        while os.path.exists(os.path.join(folder, f"{name}_v{version}{ext}")):
            version += 1
        return os.path.join(folder, f"{name}_v{version}{ext}")

    def process_file(self, file_path):
        filename = os.path.basename(file_path)
        if file_path in self.processed:
            return False
        print(f"📄 {filename}")
        category, subdir = self.find_destination(filename)
        if not category:
            print(f"  ⚠ Pas de règle")
            return False
        dest_folder = os.path.join(self.dest, category, subdir)
        os.makedirs(dest_folder, exist_ok=True)
        base, ext = os.path.splitext(filename)
        if ext.lower() == '.ifc' and IFC_AVAILABLE and self.analyze_ifc_enabled:
            self.analyze_ifc(file_path)
        target = self.get_versioned_path(dest_folder, base, ext)
        try:
            shutil.copy2(file_path, target)
            os.remove(file_path)
            self.processed.add(file_path)
            print(f"  ✓ → {os.path.basename(target)}")
            return True
        except Exception as e:
            print(f"  ✗ Erreur: {e}")
            return False

    def run(self):
        print("\n╔" + "="*60 + "╗")
        print("║  ORGANISATEUR DE FICHIERS COMPACT (Excel + IFC)         ║")
        print("╚" + "="*60 + "╝\n")
        print(f"📂 Source: {self.source}\n📁 Destination: {self.dest}\n📊 Excel: {self.excel_path}")
        if IFC_AVAILABLE and self.analyze_ifc_enabled:
            print(f"🏗️  Analyse IFC: Activée → {self.ifc_folder}")
        print()
        self.load_config()
        if not self.mapping:
            print(f"\n⚠️  Configurez {self.excel_path} et relancez le script\n")
            return
        if not os.path.exists(self.source):
            print(f"✗ Dossier source introuvable: {self.source}")
            return
        files = [os.path.join(self.source, f) for f in os.listdir(self.source)
                if os.path.isfile(os.path.join(self.source, f)) and not f.startswith('.') and not f.startswith('~')]
        print(f"📂 {len(files)} fichier(s) trouvé(s)\n")
        success = sum(1 for f in files if self.process_file(f))
        print(f"\n✓ Terminé: {success}/{len(files)} fichiers traités\n")

def main():
    organizer = FileOrganizer()
    organizer.run()

if __name__ == "__main__":
    main()

